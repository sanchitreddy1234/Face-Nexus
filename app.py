
from flask import Flask, render_template, request, redirect, session, url_for, flash
import os
import cv2
import numpy as np
from tensorflow.keras.models import load_model, Sequential  # type: ignore
from tensorflow.keras.preprocessing.image import img_to_array  # type: ignore
from tensorflow.keras.layers import Dense, Flatten, Conv2D, MaxPooling2D  # type: ignore
from tensorflow.keras.applications.inception_v3 import InceptionV3  # type: ignore
from tensorflow.keras.layers import Dense, GlobalAveragePooling2D  # type: ignore
from tensorflow.keras.models import Model  # type: ignore
from keras.layers import Input, Conv2D, GlobalAveragePooling2D, Dense, Lambda # type: ignore
from tensorflow.keras.preprocessing.image import ImageDataGenerator  # type: ignore
import pandas as pd
import tensorflow as tf
import matplotlib.pyplot as plt  # type: ignore
import io
import base64
import openpyxl  # type: ignore
from datetime import datetime
from werkzeug.utils import secure_filename
import time

app = Flask(__name__)
app.secret_key = "attendance_secret_key"

# Paths
DATASET_DIR = "face_recognition_dataset"
MODEL_PATH = "face_model.h5"
ATTENDANCE_FILE = "attendance1.xlsx"

# Load Class Indices
if os.path.exists("class_new.npy"):
    class_indices = np.load("class_new.npy", allow_pickle=True).item()
else:
    class_indices = {}

reverse_mapping = {v: k for k, v in class_indices.items()}
print(reverse_mapping)

# Define dataset and model paths
DATASET_DIR = r"face_recognition_dataset"
MODEL_PATH = r"model.h5"
if not os.path.exists(ATTENDANCE_FILE):
    wb = openpyxl.Workbook()  # Create a new workbook if file doesn't exist
    sheet = wb.active
    sheet.append(["Name"])  # Add header row for names only
    
    # Add student names from class_indices.values() to the Excel file
    for student_name in reverse_mapping.values():  # Directly using values to get student names
        sheet.append([student_name])  # Add student names without any default status
    
    wb.save(ATTENDANCE_FILE) 

# Database connection configuration
DB_CONFIG = {
    "host": "localhost",
    "user": "root",
    "password": "",
    "database": "attendence"
}

if not os.path.exists(ATTENDANCE_FILE):
    wb = openpyxl.Workbook()  # Create a new workbook if file doesn't exist
    sheet = wb.active
    sheet.append(["Name"])  # Add header row for names only
    
    # Add student names from class_indices.values() to the Excel file
    for student_name in reverse_mapping.values():  # Directly using values to get student names
        sheet.append([student_name])  # Add student names without any default status
    
    wb.save(ATTENDANCE_FILE) 


# Load data for training
def load_data(img_height, img_width):
    # Data Augmentation
    datagen = ImageDataGenerator(
        rescale=1.0 / 255,            # Normalizing the image
        validation_split=0.3,         # Split dataset for validation
        rotation_range=30,            # Randomly rotate images by 30 degrees
        width_shift_range=0.2,        # Horizontally shift images by 20% of the width
        height_shift_range=0.2,       # Vertically shift images by 20% of the height
        shear_range=0.2,              # Shear transformation with a shear intensity of 0.2
        zoom_range=0.2,               # Randomly zoom in on images by 20%
        horizontal_flip=True,         # Randomly flip images horizontally
        fill_mode='nearest'   
    )

    train_data = datagen.flow_from_directory(
        DATASET_DIR,
        target_size=(img_height, img_width),
        batch_size=32,
        class_mode="categorical",
        subset="training"
    )
    val_data = datagen.flow_from_directory(
        DATASET_DIR,
        target_size=(img_height, img_width),
        batch_size=32,
        class_mode="categorical",
        subset="validation"
    )

    return train_data, val_data






# Create the model
def create_model(input_shape, num_classes):
    # Create the Base Model
    base_model = InceptionV3(weights='imagenet', include_top=False, input_shape=input_shape)
    
    # Freeze the base model (optional)
    for layer in base_model.layers:
        layer.trainable = False
    
    # Add Custom Top Layers
    x = base_model.output
    x = GlobalAveragePooling2D()(x)
    
    predictions = Dense(num_classes, activation='softmax')(x)  # Adjust the number of classes as needed
    
    # Create the Final Model
    model = Model(inputs=base_model.input, outputs=predictions)
    
    # Compile the Model
    model.compile(optimizer='adam', loss='categorical_crossentropy', metrics=['accuracy'])
    
    #model.summary()

    return model


# @app.route("/index")
# def index():
#     return render_template("index.html")



@app.route("/")
def home():
    return render_template("index.html")

@app.route("/about")
def about():
    return render_template("about.html")

@app.route("/contact")
def contact():
    return render_template("contact.html")

# @app.route("/studentlogin")
# def studentlogin():
#     return render_template("studentlogin.html")

@app.route("/adminlogin")
def adminlogin():
    return render_template("adminlogin.html")

@app.route("/studentlogin", methods=["GET", "POST"])
def studentlogin():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        if not username or not password:
            flash("Please provide both username and password.", "error")
            return redirect(url_for("studentlogin"))

        # Validate user credentials against the database
        import pymysql

        try:
            conn = pymysql.connect(
                host="localhost",
                user="root",
                password="",
                database="attendence"
            )
            cursor = conn.cursor()
            cursor.execute(
                "SELECT username FROM user_login_info WHERE username = %s AND password = %s",
                (username, password)
            )
            user = cursor.fetchone()
            cursor.close()
            conn.close()

            if user:
                flash("Login successful!", "success")
                return redirect(url_for("analysis2", student_name=username))
            else:
                flash("Invalid username or password.", "error")
                return redirect(url_for("studentlogin"))

        except Exception as e:
            flash("Database error: " + str(e), "error")
            return redirect(url_for("studentlogin"))

    # Render the login page for GET requests
    return render_template("studentlogin.html")


@app.route("/facultylogin", methods=["GET", "POST"])
def facultylogin():
    if request.method == "POST":
        # Get username and password from the form
        email = request.form.get("name")
        password = request.form.get("password")

        try:
            import pymysql
            # Connect to the database
            conn = pymysql.connect(**DB_CONFIG)
            cursor = conn.cursor()

            # Query the database for the given username and password
            query = "SELECT * FROM faculty_login_info WHERE username = %s AND password = %s"
            cursor.execute(query, (email, password))
            result = cursor.fetchone()

            cursor.close()
            conn.close()

            # If a user is found, log in and redirect
            if result:
                session["user"] = email  # Store the user in the session
                flash("Login successful!", "success")
                return redirect(url_for("analysis1"))
            else:
                flash("Invalid username or password. Please try again.", "error")
        except Exception as e:
            flash(f"Database error: {str(e)}", "error")

    return render_template("facultylogin.html")

@app.route("/logout")
def logout():
    session.pop("user", None)  # Log out the user
    flash("You have been logged out.", "info")
    return redirect(url_for("facultylogin"))


@app.route("/admin_base")
def admin_base():
    return render_template("base1.html")
# Train the model
@app.route("/train", methods=["GET", "POST"])
def train():
    if request.method == "POST":

        train_data, val_data = load_data(256, 256)
        num_classes = len(train_data.class_indices)
        import numpy as np
        np.save('class_new.npy', train_data.class_indices)

        model = create_model((256, 256, 3), num_classes)
        model.fit(train_data, validation_data=val_data, epochs=25)

        # Save the model
        model.save(MODEL_PATH)
        flash("Model trained successfully!", "success")
        return redirect(url_for("train"))

    return render_template("train.html")



@app.route("/registerfaculty", methods=["GET", "POST"])
def faculty_register():
    if request.method == "POST":
        user_name = request.form.get("name")
        password = request.form.get("password")  # Get password from the form
        
        if not user_name or not password:
            flash("Please provide both username and password", "error")
            return redirect(url_for("register"))
        
        # Save user information in the database
        import pymysql
        
        try:
            conn = pymysql.connect(
                host="localhost",
                user="root",
                password="",
                database="attendence"
            )
            cursor = conn.cursor()
            cursor.execute("INSERT INTO faculty_login_info (username, password) VALUES (%s, %s)", (user_name, password))
            conn.commit()
            cursor.close()
            conn.close()
            
        except Exception as e:
            flash("Database error: " + str(e), "error")
            return redirect(url_for("registerfaculty",))
    return render_template("facultyregistrations.html",value="registration success")


@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        user_name = request.form.get("username")
        password = request.form.get("password")  # Get password from the form
        
        if not user_name or not password:
            flash("Please provide both username and password", "error")
            return redirect(url_for("register"))
        
        # Save user information in the database
        import pymysql
        
        try:
            conn = pymysql.connect(
                host="localhost",
                user="root",
                password="",
                database="attendence"
            )
            cursor = conn.cursor()
            cursor.execute("INSERT INTO user_login_info (username, password) VALUES (%s, %s)", (user_name, password))
            conn.commit()
            cursor.close()
            conn.close()

        except Exception as e:
            flash("Database error: " + str(e), "error")
            return redirect(url_for("register"))

        # Create a directory for storing user's face images
        save_dir = os.path.join(DATASET_DIR, user_name)
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        # Capture 100 face images
        cap = cv2.VideoCapture(0)
        count = 0
        while count < 500:  # Capture 100 images
            ret, frame = cap.read()
            if ret:
                face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                faces = face_cascade.detectMultiScale(gray, 1.3, 5)

                for (x, y, w, h) in faces:
                    # Add padding around the face
                    padding = 30
                    x = max(0, x - padding)
                    y = max(0, y - padding) 
                    w = min(frame.shape[1] - x, w + 2*padding)
                    h = min(frame.shape[0] - y, h + 2*padding)
                    
                    # Crop and save the full color image with padding
                    face_img = frame[y:y+h, x:x+w]
                    img_path = os.path.join(save_dir, f"{user_name}_{count}.jpg")
                    cv2.imwrite(img_path, face_img)
                    count += 1
                
                # Log progress to console
                print(f"Captured {count}/500 images")
            cv2.imshow("Marking Attendance", frame)
            if cv2.waitKey(1) & 0xFF == ord('q') or count >= 500:
                break

        cap.release()
        cv2.destroyAllWindows()
        flash(f"User {user_name} registered successfully!", "success")
        return redirect(url_for("admin_base"))

    return render_template("register.html")

@app.route("/register1", methods=["GET", "POST"])
def register1():
    if request.method == "POST":
        user_name = request.form.get("username")
        password = request.form.get("password")  # Get password from the form
        
        if not user_name or not password:
            flash("Please provide both username and password", "error")
            return redirect(url_for("register"))
        
        # Save user information in the database
        import pymysql
        
        try:
            conn = pymysql.connect(
                host="localhost",
                user="root",
                password="",
                database="attendence"
            )
            cursor = conn.cursor()
            cursor.execute("INSERT INTO user_login_info (username, password) VALUES (%s, %s)", (user_name, password))
            conn.commit()
            cursor.close()
            conn.close()
        except Exception as e:
            flash("Database error: " + str(e), "error")
            return redirect(url_for("register"))

        # Create a directory for storing user's face images
        save_dir = os.path.join(DATASET_DIR, user_name)
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        # Capture 100 face images
        cap = cv2.VideoCapture(0)
        count = 0
        while count < 500:  # Capture 100 images
            ret, frame = cap.read()
            if ret:
                face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
                gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                faces = face_cascade.detectMultiScale(gray, 1.3, 5)

                for (x, y, w, h) in faces:
                    # Crop the grayscale face
                    face_img_gray = gray[y:y+h, x:x+w]
                    img_path = os.path.join(save_dir, f"{user_name}_{count}.jpg")
                    # Save the grayscale image
                    cv2.imwrite(img_path, face_img_gray)
                    count += 1

                #cv2.imshow("Registering User", gray)  # Display the grayscale frame

            if cv2.waitKey(1) & 0xFF == ord('q') or count >= 500:
                break

        cap.release()
        cv2.destroyAllWindows()
        flash(f"User {user_name} registered successfully!", "success")
        return redirect(url_for("admin_base"))

    return render_template("register.html")



@app.route("/mark_attendance1", methods=["GET", "POST"])
def mark_attendance1():
    if not os.path.exists(DATASET_DIR):
        flash("Dataset directory not found!", "error")
        return redirect(url_for("home"))

    attendance = {}
    cap = cv2.VideoCapture(0)
    
    # Initialize face detector
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
    
    # Load all student images
    student_images = {}
    for student_name in os.listdir(DATASET_DIR):
        student_dir = os.path.join(DATASET_DIR, student_name)
        if os.path.isdir(student_dir):
            # Get the first image of each student
            for img_name in os.listdir(student_dir):
                if img_name.endswith(('.jpg', '.jpeg', '.png')):
                    img_path = os.path.join(student_dir, img_name)
                    student_images[student_name] = img_path
                    break

    last_capture_time = time.time()
    capture_interval = 10  # Capture every 10 seconds
    last_recognized = {}  # Store last recognition time for each person
    current_best_match = None  # Store current best match
    current_best_similarity = 0  # Store current best similarity

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        current_time = time.time()
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        # Only process faces every 10 seconds
        if current_time - last_capture_time >= capture_interval:
            for (x, y, w, h) in faces:
                try:
                    # Get face ROI from current frame
                    face_img = frame[y:y+h, x:x+w]
                    
                    # Resize face to standard size
                    face_img = cv2.resize(face_img, (256, 256))
                    
                    best_match = None
                    best_similarity = 0
                    
                    # Compare with all student images
                    for student_name, stored_img_path in student_images.items():
                        try:
                            # Read stored image
                            stored_img = cv2.imread(stored_img_path)
                            if stored_img is None:
                                print(f"Could not read image: {stored_img_path}")
                                continue
                            
                            # Detect face in stored image
                            stored_gray = cv2.cvtColor(stored_img, cv2.COLOR_BGR2GRAY)
                            stored_faces = face_cascade.detectMultiScale(stored_gray, 1.3, 5)
                            
                            if len(stored_faces) == 0:
                                continue
                                
                            # Get the first face from stored image
                            sx, sy, sw, sh = stored_faces[0]
                            stored_face = stored_img[sy:sy+sh, sx:sx+sw]
                            stored_face = cv2.resize(stored_face, (256, 256))
                            
                            # Convert both faces to grayscale
                            face_gray = cv2.cvtColor(face_img, cv2.COLOR_BGR2GRAY)
                            stored_face_gray = cv2.cvtColor(stored_face, cv2.COLOR_BGR2GRAY)
                            
                            # Calculate similarity using template matching
                            result = cv2.matchTemplate(face_gray, stored_face_gray, cv2.TM_CCOEFF_NORMED)
                            similarity = np.max(result)
                            
                            print(f"Comparing with {student_name}: similarity = {similarity:.2f}")
                            
                            if similarity > best_similarity:
                                best_similarity = similarity
                                best_match = student_name
                                
                        except Exception as e:
                            print(f"Error comparing with {student_name}: {str(e)}")
                            continue
                    
                    # Update current best match
                    if best_match and best_similarity > 0.1:  # Threshold for similarity
                        current_best_match = best_match
                        current_best_similarity = best_similarity
                        
                        # Update attendance if enough time has passed since last recognition
                        if current_best_match not in last_recognized or \
                           current_time - last_recognized[current_best_match] >= capture_interval:
                            attendance[current_best_match] = "Present"
                            last_recognized[current_best_match] = current_time
                            print(f"Recognized {current_best_match} with similarity {current_best_similarity:.2f}")
                    
                except Exception as e:
                    print(f"Error processing face: {str(e)}")
                    continue
            
            last_capture_time = current_time

        # Always draw rectangles and labels on the frame
        for (x, y, w, h) in faces:
            try:
                # Draw rectangle and label based on current best match
                if current_best_match and current_best_similarity > 0.5:
                    color = (0, 255, 0) if current_best_similarity > 0.7 else (0, 255, 255)
                    cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                    label = f"{current_best_match} ({current_best_similarity:.2f})"
                    cv2.putText(frame, label, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)
                else:
                    cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 255), 2)
                    cv2.putText(frame, "Unknown", (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
            except Exception as e:
                print(f"Error drawing rectangle: {str(e)}")
                continue

        # Display the frame
        cv2.imshow("Marking Attendance", frame)
        
        # Check for 'q' key to quit
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

    # Save attendance to Excel
    if not os.path.exists(ATTENDANCE_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Name", "Status", "Timestamp"])  # Add header row
        wb.save(ATTENDANCE_FILE)

    wb = openpyxl.load_workbook(ATTENDANCE_FILE)
    sheet = wb.active
    existing_names = {row[0].value for row in sheet.iter_rows(min_row=2, max_col=1)}

    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for user_name, status in attendance.items():
        if user_name not in existing_names:
            sheet.append([user_name, status, current_time])
    wb.save(ATTENDANCE_FILE)

    flash("Attendance marked successfully!", "success")
    return redirect(url_for("admin_base"))

# import re
# from datetime import datetime
# @app.route("/mark_attendance", methods=["GET", "POST"])
# def mark_attendance():
#     if not os.path.exists(MODEL_PATH):
#         flash("Model not found! Train the CNN first.", "error")
#         return redirect(url_for("admin_base"))

#     # Load the pre-trained CNN model
#     model = load_model(MODEL_PATH)
#     attendance = {}
#     cap = cv2.VideoCapture(0)

#     # Set a threshold for prediction confidence
#     CONFIDENCE_THRESHOLD = 0.85  # Adjust as needed

#     while True:
#         ret, frame = cap.read()
#         if not ret:
#             break

#         face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
#         gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
#         faces = face_cascade.detectMultiScale(gray, 1.3, 5)

#         for (x, y, w, h) in faces:
#             # Preprocess the face for the model
#             face_img = frame[y:y+h, x:x+w]
#             face_img = cv2.cvtColor(face_img, cv2.COLOR_BGR2RGB)  # Convert to RGB if required
#             face_img = cv2.resize(face_img, (100, 100))           # Resize to model's input size
#             face_img = face_img / 255.0                           # Normalize
#             face_img = np.expand_dims(face_img, axis=0)           # Add batch dimension

#             # Make a prediction
#             prediction = model.predict(face_img)
#             predicted_class = np.argmax(prediction)
#             confidence = prediction[0][predicted_class]
#             print(predicted_class)

#             # Apply the threshold
#             if confidence >= CONFIDENCE_THRESHOLD:
#                 class_name = reverse_mapping.get(predicted_class, "Unknown")
                
#                 attendance[class_name] = "Present"

#                 # Draw a rectangle and label on the frame
#                 cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 0, 0), 2)
#                 label = f"{class_name} ({confidence*100:.1f}%)"
#                 cv2.putText(frame, label, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 2)

#         # Display the frame
#         cv2.imshow("Marking Attendance", frame)
#         if cv2.waitKey(1) & 0xFF == ord('q'):
#             break

#     cap.release()
#     cv2.destroyAllWindows()

#     # Save attendance to Excel
#     if not os.path.exists(ATTENDANCE_FILE):
#         wb = openpyxl.Workbook()
#         sheet = wb.active
#         sheet.append(["Name", "Status"])  # Add header row
#         wb.save(ATTENDANCE_FILE)

#     wb = openpyxl.load_workbook(ATTENDANCE_FILE)
#     sheet = wb.active

#     current_date = datetime.now().strftime("%Y-%m-%d")  # Format as "YYYY-MM-DD"
#     # Check if the column for the current date exists
#     header = [cell.value for cell in sheet[1]]  # Get the first row (header)
    
#     if current_date not in header:
#         header.append(current_date)  # Add current date to the header list
#         # Re-write the updated header back to the sheet
#         for col_idx, value in enumerate(header, start=1):  # `start=1` to begin from column 1
#             sheet.cell(row=1, column=col_idx, value=value)
#     # Find the index of the current date in the header
#     date_column_idx = header.index(current_date) + 1  # +1 to convert zero-based index to Excel's column index
#     # Iterate over each row (students)
#     for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column):
#         name = row[0].value  # First column should be "Name"
#         if name in attendance:
#             try:
#                 # Ensure the row has enough columns to write the attendance
#                 if len(row) < date_column_idx:
#                     for i in range(len(row), date_column_idx):
#                         sheet.cell(row=row[0].row, column=i + 1, value="")
#                 row[date_column_idx - 1].value = attendance[name]  # Mark attendance in the corresponding date column
#             except IndexError:
#                 # In case there is an issue with row access, log the error
#                 flash(f"Error marking attendance for {name}.", "error")

#     wb.save(ATTENDANCE_FILE)

#     flash("Attendance marked successfully!", "success")
#     return redirect(url_for("admin_base"))

import os
import cv2
import numpy as np
import openpyxl
from datetime import datetime
from flask import flash, redirect, url_for
from tensorflow.keras.models import load_model

@app.route("/mark_attendance", methods=["GET", "POST"])
def mark_attendance():
    if not os.path.exists(MODEL_PATH):
        flash("Model not found! Train the CNN first.", "error")
        return redirect(url_for("admin_base"))

    # Load the pre-trained CNN model
    model = load_model(MODEL_PATH)
    attendance = {}  # Final attendance record

    # Detection tracking
    detection_counter = {}  # To count consecutive detections
    REQUIRED_CONSECUTIVE_DETECTIONS = 40  # Number of frames required for confirmation
    RESET_AFTER_FRAMES = 10  # Frames after which detection count resets if not seen
    CONFIDENCE_THRESHOLD = 0.90  # Confidence threshold for detection

    cap = cv2.VideoCapture(0)
    face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)

        current_detected_names = set()  # Names detected in this frame

        for (x, y, w, h) in faces:
            # Preprocess the face
            face_img = frame[y:y + h, x:x + w]
            face_img = cv2.cvtColor(face_img, cv2.COLOR_BGR2RGB)
            face_img = cv2.resize(face_img, (256, 256))
            face_img = face_img / 255.0
            face_img = np.expand_dims(face_img, axis=0)

            # Prediction
            prediction = model.predict(face_img)
            predicted_class = np.argmax(prediction)
            confidence = prediction[0][predicted_class]
            class_name = reverse_mapping.get(predicted_class, "Unknown")

            if confidence >= CONFIDENCE_THRESHOLD:
                current_detected_names.add(class_name)

                # Update consecutive detection counter
                if class_name not in detection_counter:
                    detection_counter[class_name] = {'count': 1, 'marked': False, 'missed': 0}
                else:
                    if not detection_counter[class_name]['marked']:
                        detection_counter[class_name]['count'] += 1
                        detection_counter[class_name]['missed'] = 0  # Reset missed when detected again

                # Check and mark attendance if threshold reached
                if (detection_counter[class_name]['count'] >= REQUIRED_CONSECUTIVE_DETECTIONS and
                        not detection_counter[class_name]['marked']):
                    attendance[class_name] = "Present"
                    detection_counter[class_name]['marked'] = True
                    print(f"[INFO] Attendance marked for {class_name}")

                # Draw rectangle and label
                count_display = detection_counter[class_name]['count']
                label = f"{class_name} ({confidence * 100:.1f}%) {count_display}/{REQUIRED_CONSECUTIVE_DETECTIONS}"
                if detection_counter[class_name]['marked']:
                    label += " ✓"

                cv2.rectangle(frame, (x, y), (x + w, y + h), (255, 0, 0), 2)
                cv2.putText(frame, label, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 0, 0), 2)

        # Decay counters for people not detected this frame
        for name in detection_counter.keys():
            if name not in current_detected_names and not detection_counter[name]['marked']:
                detection_counter[name]['missed'] += 1
                if detection_counter[name]['missed'] >= RESET_AFTER_FRAMES:
                    detection_counter[name]['count'] = 0  # Reset detection count
                    detection_counter[name]['missed'] = 0  # Reset missed counter

        # Display frame
        cv2.imshow("Marking Attendance", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

    # ---------------------- Save attendance to Excel ----------------------

    if not os.path.exists(ATTENDANCE_FILE):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Name", "Status"])  # Add header
        wb.save(ATTENDANCE_FILE)

    wb = openpyxl.load_workbook(ATTENDANCE_FILE)
    sheet = wb.active

    current_date = datetime.now().strftime("%Y-%m-%d")
    header = [cell.value for cell in sheet[1]]  # Header row

    if current_date not in header:
        header.append(current_date)  # Add current date to header
        for col_idx, value in enumerate(header, start=1):
            sheet.cell(row=1, column=col_idx, value=value)

    date_column_idx = header.index(current_date) + 1  # Column index for today's date

    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column):
        name = row[0].value
        if name in attendance:
            try:
                if len(row) < date_column_idx:
                    for i in range(len(row), date_column_idx):
                        sheet.cell(row=row[0].row, column=i + 1, value="")
                row[date_column_idx - 1].value = attendance[name]  # Mark attendance
            except IndexError:
                flash(f"Error marking attendance for {name}.", "error")

    wb.save(ATTENDANCE_FILE)
    flash("Attendance marked successfully!", "success")
    return redirect(url_for("admin_base"))


import re
import seaborn as sns
@app.route("/analysis1")
def analysis1():
    if not os.path.exists(ATTENDANCE_FILE):
        flash("No attendance data found! ", "error")
        return redirect(url_for("admin_base"))

    # Read the attendance data without specifying column names
    df = pd.read_excel(ATTENDANCE_FILE)

    # Check if the dataframe has columns "User Name", "Status", "Timestamp"
    if "Name" not in df.columns :
        flash("Attendance data is malformed!", "error")
        return redirect(url_for("admin_base"))

    # Convert all column names to strings and check for date columns
    df.columns = df.columns.astype(str)
    date_pattern = r'^\d{4}-\d{2}-\d{2}$'
    date_columns = [col for col in df.columns if re.match(date_pattern, col)]

    if not date_columns:
        flash("No date columns found in the attendance data.", "error")
        return redirect(url_for("home"))

    # Calculate the number of days being analyzed
    num_days = len(date_columns)
    flash(f"Analyzing attendance for {num_days} days.", "info")

    # Prepare attendance summary (days present/absent)
    attendance_summary = df.set_index('Name').reindex(columns=date_columns).applymap(lambda x: 1 if 'Present' in str(x) else 0)
    attendance_summary['Total Present'] = attendance_summary.sum(axis=1)
    attendance_summary['Total Absent'] = num_days - attendance_summary['Total Present']
    attendance_summary['Attendance Percentage'] = (attendance_summary['Total Present'] / num_days) * 100

    # Plot attendance summary for percentage and presence counts
    plots = []
    for date_column in date_columns:
        # Group by user and count "Present" status for each date column
        daily_summary = df.groupby("Name")[date_column].apply(
            lambda x: x.str.contains('Present', case=False).sum() if x.dtype == 'object' else 0
        )

        # Create a more appealing bar plot using Seaborn
        plt.figure(figsize=(10, 5))
        sns.barplot(x=daily_summary.index, y=daily_summary.values, palette="coolwarm")
        plt.title(f"Attendance Summary for {date_column}", fontsize=16)
        plt.xlabel("User Name", fontsize=12)
        plt.ylabel("Days Present", fontsize=12)
        plt.xticks(rotation=45)
        plt.tight_layout()

        # Save the plot as a base64-encoded string
        img = io.BytesIO()
        plt.savefig(img, format="png")
        img.seek(0)
        plot_url = base64.b64encode(img.getvalue()).decode()
        plt.close()

        # Append the plot URL and the date column to the list
        plots.append({"plot_url": plot_url, "date_column": date_column})

    # Total attendance percentage plot
    plt.figure(figsize=(10, 5))
    sns.barplot(x=attendance_summary.index, y=attendance_summary['Attendance Percentage'], palette="viridis")
    plt.title(f"Attendance Percentage for Each Student", fontsize=16)
    plt.xlabel("User Name", fontsize=12)
    plt.ylabel("Attendance Percentage (%)", fontsize=12)
    plt.xticks(rotation=45)
    plt.tight_layout()
    img = io.BytesIO()
    plt.savefig(img, format="png")
    img.seek(0)
    plot_url = base64.b64encode(img.getvalue()).decode()
    plt.close()

    # Add the overall percentage plot
    plots.append({"plot_url": plot_url, "date_column": "Total Attendance Percentage"})

    # Heatmap of attendance data
    plt.figure(figsize=(10, 6))
    sns.heatmap(attendance_summary.drop(columns=['Total Present', 'Total Absent', 'Attendance Percentage']),
                annot=True, cmap="coolwarm", cbar=False, linewidths=0.5)
    plt.title("Heatmap of Attendance Across Dates", fontsize=16)
    plt.tight_layout()
    img = io.BytesIO()
    plt.savefig(img, format="png")
    img.seek(0)
    plot_url = base64.b64encode(img.getvalue()).decode()
    plt.close()

    plots.append({"plot_url": plot_url, "date_column": "Attendance Heatmap"})

    # Render the analysis page with all plots
    return render_template("analysis1.html", plots=plots, num_days=num_days, attendance_summary=attendance_summary)

import seaborn as sns
import matplotlib.pyplot as plt
import re
@app.route("/analysis")
def analysis():
    if not os.path.exists(ATTENDANCE_FILE):
        flash("No attendance data found! ", "error")
        return redirect(url_for("admin_base"))

    # Read the attendance data without specifying column names
    df = pd.read_excel(ATTENDANCE_FILE)

    # Check if the dataframe has columns "User Name", "Status", "Timestamp"
    if "Name" not in df.columns :
        flash("Attendance data is malformed!", "error")
        return redirect(url_for("admin_base"))

    # Convert all column names to strings and check for date columns
    df.columns = df.columns.astype(str)
    date_pattern = r'^\d{4}-\d{2}-\d{2}$'
    date_columns = [col for col in df.columns if re.match(date_pattern, col)]

    if not date_columns:
        flash("No date columns found in the attendance data.", "error")
        return redirect(url_for("home"))

    # Calculate the number of days being analyzed
    num_days = len(date_columns)
    flash(f"Analyzing attendance for {num_days} days.", "info")

    # Prepare attendance summary (days present/absent)
    attendance_summary = df.set_index('Name').reindex(columns=date_columns).applymap(lambda x: 1 if 'Present' in str(x) else 0)
    attendance_summary['Total Present'] = attendance_summary.sum(axis=1)
    attendance_summary['Total Absent'] = num_days - attendance_summary['Total Present']
    attendance_summary['Attendance Percentage'] = (attendance_summary['Total Present'] / num_days) * 100

    # Plot attendance summary for percentage and presence counts
    plots = []
    for date_column in date_columns:
        # Group by user and count "Present" status for each date column
        daily_summary = df.groupby("Name")[date_column].apply(
            lambda x: x.str.contains('Present', case=False).sum() if x.dtype == 'object' else 0
        )

        # Create a more appealing bar plot using Seaborn
        plt.figure(figsize=(10, 5))
        sns.barplot(x=daily_summary.index, y=daily_summary.values, palette="coolwarm")
        plt.title(f"Attendance Summary for {date_column}", fontsize=16)
        plt.xlabel("User Name", fontsize=12)
        plt.ylabel("Days Present", fontsize=12)
        plt.xticks(rotation=45)
        plt.tight_layout()

        # Save the plot as a base64-encoded string
        img = io.BytesIO()
        plt.savefig(img, format="png")
        img.seek(0)
        plot_url = base64.b64encode(img.getvalue()).decode()
        plt.close()

        # Append the plot URL and the date column to the list
        plots.append({"plot_url": plot_url, "date_column": date_column})

    # Total attendance percentage plot
    plt.figure(figsize=(10, 5))
    sns.barplot(x=attendance_summary.index, y=attendance_summary['Attendance Percentage'], palette="viridis")
    plt.title(f"Attendance Percentage for Each Student", fontsize=16)
    plt.xlabel("User Name", fontsize=12)
    plt.ylabel("Attendance Percentage (%)", fontsize=12)
    plt.xticks(rotation=45)
    plt.tight_layout()
    img = io.BytesIO()
    plt.savefig(img, format="png")
    img.seek(0)
    plot_url = base64.b64encode(img.getvalue()).decode()
    plt.close()

    # Add the overall percentage plot
    plots.append({"plot_url": plot_url, "date_column": "Total Attendance Percentage"})

    # Heatmap of attendance data
    plt.figure(figsize=(10, 6))
    sns.heatmap(attendance_summary.drop(columns=['Total Present', 'Total Absent', 'Attendance Percentage']),
                annot=True, cmap="coolwarm", cbar=False, linewidths=0.5)
    plt.title("Heatmap of Attendance Across Dates", fontsize=16)
    plt.tight_layout()
    img = io.BytesIO()
    plt.savefig(img, format="png")
    img.seek(0)
    plot_url = base64.b64encode(img.getvalue()).decode()
    plt.close()

    plots.append({"plot_url": plot_url, "date_column": "Attendance Heatmap"})

    # Render the analysis page with all plots
    return render_template("analysis.html", plots=plots, num_days=num_days, attendance_summary=attendance_summary)

@app.route("/analysis1/<student_name>")
def analysis2(student_name):
    if not os.path.exists(ATTENDANCE_FILE):
        flash("No attendance data found! ", "error")
        return redirect(url_for("admin_base"))

    # Read the attendance data without specifying column names
    df = pd.read_excel(ATTENDANCE_FILE)

    # Check if the dataframe has columns "Name", "Status", "Timestamp"
    if "Name" not in df.columns:
        flash("Attendance data is malformed!", "error")
        return redirect(url_for("admin_base"))

    # Convert all column names to strings and check for date columns
    df.columns = df.columns.astype(str)
    date_pattern = r'^\d{4}-\d{2}-\d{2}$'
    date_columns = [col for col in df.columns if re.match(date_pattern, col)]

    if not date_columns:
        flash("No date columns found in the attendance data.", "error")
        return redirect(url_for("home"))

    # Check if the student exists in the dataset
    if student_name not in df['Name'].values:
        flash(f"Student {student_name} not found in the attendance data!", "error")
        return redirect(url_for("home"))

    # Filter the dataframe for the specific student
    student_data = df[df['Name'] == student_name].set_index('Name')

    # Calculate the number of days being analyzed
    num_days = len(date_columns)
    flash(f"Analyzing attendance for {num_days} days for {student_name}.", "info")

    # Prepare attendance summary for the specific student
    attendance_summary = student_data.reindex(columns=date_columns).applymap(
        lambda x: 1 if 'Present' in str(x) else 0
    )
    attendance_summary['Total Present'] = attendance_summary.sum(axis=1)
    attendance_summary['Total Absent'] = num_days - attendance_summary['Total Present']
    attendance_summary['Attendance Percentage'] = (attendance_summary['Total Present'] / num_days) * 100

    # Generate plots for the specific student
    plots = []

    # Daily attendance plot
    daily_summary = attendance_summary.iloc[0]
    plt.figure(figsize=(10, 5))
    sns.barplot(x=date_columns, y=daily_summary[date_columns].values, palette="coolwarm")
    plt.title(f"Daily Attendance for {student_name}", fontsize=16)
    plt.xlabel("Dates", fontsize=12)
    plt.ylabel("Attendance (1=Present, 0=Absent)", fontsize=12)
    plt.xticks(rotation=45)
    plt.tight_layout()

    # Save the plot as a base64-encoded string
    img = io.BytesIO()
    plt.savefig(img, format="png")
    img.seek(0)
    plot_url = base64.b64encode(img.getvalue()).decode()
    plt.close()

    # Append the plot URL and the date column to the list
    plots.append({"plot_url": plot_url, "date_column": "Daily Attendance"})

    # Attendance percentage plot
    plt.figure(figsize=(10, 5))
    sns.barplot(x=['Total Present', 'Total Absent'], y=[
                attendance_summary.iloc[0]['Total Present'], attendance_summary.iloc[0]['Total Absent']], palette="viridis")
    plt.title(f"Attendance Summary for {student_name}", fontsize=16)
    plt.xlabel("Attendance Type", fontsize=12)
    plt.ylabel("Count", fontsize=12)
    plt.tight_layout()
    img = io.BytesIO()
    plt.savefig(img, format="png")
    img.seek(0)
    plot_url = base64.b64encode(img.getvalue()).decode()
    plt.close()

    plots.append({"plot_url": plot_url, "date_column": "Attendance Summary"})

    # Render the analysis page with the specific student's plots
    return render_template("analysis1.html", plots=plots, num_days=num_days, attendance_summary=attendance_summary)



if __name__ == "__main__":
    app.run(debug=False)
